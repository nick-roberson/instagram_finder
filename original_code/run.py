import instaloader
import os
from PIL import Image
import xlrd
from xlrd import open_workbook

# from newsapi import NewsApiClient
from datetime import timedelta
from datetime import datetime
import xlwt

# from xlwt import Workbook
import xlsxwriter
import shutil
import time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import math
from math import log
import json
from statistics import mode
from statistics import mean
import matplotlib.pyplot as plt

# from mpl_finance import candlestick_ohlc
import pandas as pd
import matplotlib.dates as mpl_dates

# from mpl_finance import candlestick2_ohlc
import time
import numpy as np
from numpy import arange
import pandas as pd
import math
from bs4 import BeautifulSoup
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import ast
from datetime import date
from scipy.stats.stats import pearsonr
from itertools import combinations
from math import comb
from scipy import stats
import psutil
import sys  # to access the system
import cv2

# from requests_html import HTMLSession
from bs4 import SoupStrainer

# import org.openqa.selenium.Keys

universal_username = input(
    "LOGIN FIRST IN THE TERMINAL WINDOW WITH 'instaloader --login [username]' (Write username you logge intoto Confirm): "
)
wb = load_workbook(filename="desktop/IG_Test.xlsx")
sheet_ranges = wb["Sheet1"]
finaldict = {}


def IG_usernames():
    link_list = []
    username_list = []
    peak = len(sheet_ranges["A"]) + 1
    count = 1
    for item in sheet_ranges["A"]:
        if count < peak:
            val = sheet_ranges["A%d" % (count)].value
            if "instagram" in str(val):
                link_list.append(val)
                stop = str(val).find("/?")
                start = str(val).find(".com/") + 5
                username = val[start:stop]
                username_list.append(username)
                print(link_list)
                print(username_list)
                count += 1
            else:
                print("REJECTED", val)
                count += 1
    count = 0
    for item in username_list:
        finaldict[item] = [link_list[count]]
        count += 1
    return finaldict


def file_cleanup(og_file, imagery=1):
    if imagery == 1:
        print(og_file)
        z = os.listdir(str(og_file))
        print("OG", z)
        print(" ")
        for file in z:
            if "jpg" not in file:  # and "mp4" not in file:
                os.remove(og_file + "/%s" % (file))
        print("final file ", os.path.abspath(str(og_file)))
    new_file = os.path.abspath(str(og_file))
    return new_file


def imagery_approval(photo_path):
    print("Photo_path", photo_path)
    photo_file = os.listdir(photo_path)
    print(photo_file)
    location_x = 0
    location_y = 0
    base = cv2.imread(photo_path + "/%s" % (photo_file[0]), cv2.IMREAD_ANYCOLOR)
    # for item in photo_file[1:]:
    count = 1
    img1 = cv2.imread(photo_path + "/%s" % (photo_file[0]), cv2.IMREAD_ANYCOLOR)
    IMAGE_WIDTH = img1.shape[0]
    IMAGE_HEIGHT = img1.shape[1]
    img1 = cv2.resize(img1, (IMAGE_WIDTH, IMAGE_WIDTH))
    print(IMAGE_WIDTH, IMAGE_WIDTH)
    while count < len(photo_file):
        img2 = cv2.imread(photo_path + "/%s" % (photo_file[count]), cv2.IMREAD_ANYCOLOR)
        img2 = cv2.resize(img2, (IMAGE_WIDTH, IMAGE_WIDTH))
        img2 = cv2.putText(
            img=np.copy(img2),
            text="%s" % (count + 1),
            org=(200, 200),
            fontFace=3,
            fontScale=8,
            color=(255, 255, 255),
            thickness=15,
        )
        if count == 1:
            hori = np.concatenate((img1, img2), axis=1)
            verti = np.concatenate((img1, img2), axis=0)
        else:
            try:
                verti = np.concatenate((verti, img2), axis=0)
                hori = np.concatenate((hori, img2), axis=1)
            except:
                print("SKIPPED", count)
                count += 1
                continue
        # cv2.imshow('HORIZONTAL', hori)
        # cv2.waitKey(400)
        count += 1
        # print (count)
        # cv2.destroyAllWindows()
    cv2.imshow("HORIZONTAL", hori)
    cv2.waitKey(400)
    approval_lst = []
    check_out = 0
    while check_out == 0:
        approval = input("Type Number, Type 99 to END")
        if approval == "99":
            cv2.destroyAllWindows()
            check_out = 99
        else:
            approval_lst.append(int(approval))
    count = 0
    for photo in photo_file:
        if count in approval_lst:
            continue
        else:
            os.remove(photo_path + "/%s" % (photo))
        count += 1
    print("FINAL", os.listdir(photo_path))
    return photo_path


"""
		#img = Image.open(photo_path+"/%s" % (photo))
		#img.show()
		img = cv2.imread(photo_path+"/%s" % (photo), cv2.IMREAD_ANYCOLOR)
		cv2.imshow("Sheep", img)
		cv2.moveWindow("Sheep", location_x, location_y) 
		cv2.waitKey(400)
		approval = (input("1=Yes"))
		location_x += 50
		if location_x > 200:
			location_y += 50
		#cv2.destroyAllWindows()
		if approval == "1":
		#	continue
		else:
		#	os.remove(photo_path+'/%s' % (photo))"""


def ethnicity():
    eth_list = [
        "White",
        "Hispanic",
        "Black/African American",
        "Asian",
        "Native American",
        "Mixed/Unknown",
    ]
    count = 1
    for item in eth_list:
        if count < 7:
            print(count, eth_list[count - 1])
            count += 1
    eth_num = int(input("What Ethnicity"))
    eth = eth_list[eth_num - 1]
    return eth


def printer(finaldict, platform):
    wb = load_workbook(filename="desktop/PR_Test.xlsx")
    sheet_ranges = wb["Sheet1"]
    count = len(sheet_ranges["A"])
    if platform == "IG":
        count = 1
        for item in finaldict.keys():
            sheet_ranges["A%d" % count] = item
            sheet_ranges["B%d" % count] = finaldict[item][0]
            sheet_ranges["C%d" % count] = finaldict[item][1]
            sheet_ranges["D%d" % count] = finaldict[item][2]
            sheet_ranges["E%d" % count] = finaldict[item][3]
            sheet_ranges["F%d" % count] = finaldict[item][4]
            sheet_ranges["F%d" % count] = finaldict[item][5]
            count += 1
        today = str(date.today())
        name = input("Name of File (No Spaces)")
    if platform == "TT":
        wb = load_workbook(filename="desktop/PR_Test.xlsx")
        sheet_ranges = wb["Sheet1"]
        count = len(sheet_ranges["A"])
        for item in finaldict.keys():
            sheet_ranges["A%d" % count] = item
            sheet_ranges["B%d" % count] = finaldict[item][0]
            sheet_ranges["C%d" % count] = finaldict[item][1]
            sheet_ranges["D%d" % count] = finaldict[item][2]
            sheet_ranges["E%d" % count] = finaldict[item][3]
            count += 1
        today = str(date.today())
        name = input("Name of File (No Spaces)")
    wb.save(filename="desktop/Fohr_Outputs/%s_%s.xlsx" % (name, today))


def data_pull(finaldict, inputs=1):
    ig = instaloader.Instaloader()
    ig.load_session_from_file(universal_username)
    for item in finaldict.keys():
        usrname = item
        profile = instaloader.Profile.from_username(ig.context, usrname)
        print("-----------------------------------")
        print("-----------------------------------")
        print(profile.username)
        followers = profile.followers
        print(followers)
        print(profile.biography)
        if inputs == 1:
            name = input("Name? ")
            finaldict[item].append(name)
        else:
            finaldict[item].append("ENTER NAME")
        finaldict[item].append(followers)
        # instaloader.Instaloader().download_profile(usrname,profile_pic_only=True)
        print("GETTING POSTS")
        posts_list = profile.get_posts()
        """posts_sorted_by_likes = sorted(profile.get_posts(),
		                               key=lambda p: p.likes + p.comments,
		                               reverse=True)
		#for post in posts_sorted_by_likes:"""
        print("Calculating engagment rate")

        count = 0
        engagement_count = 0
        eng_dict = {}
        for post in posts_list:
            count += 1
            if count < 50:
                count += 1
                total = post.likes + post.comments
                engagement_count += total
                eng_dict[count] = total
                # print ("TYPE",post.typename)
            else:
                engagement_rate = 100 * (engagement_count / int(followers))
                finaldict[item].append(engagement_rate)
                print(engagement_rate, engagement_count, followers)
                break
        count = 0
        temp = {k: v for k, v in sorted(eng_dict.items(), key=lambda item: item[1])}
        top_list = list(temp.keys())[0:5]
        print("Downloading Top Posts", top_list)
        for post in posts_list:
            if count in top_list:
                print(count)
                ig.download_post(post, profile.username)
            if count >= 50:
                break
            count += 1
            print("Post#", count)
        print("DOWNLOADED!")
        print(finaldict)
    if inputs == 1:
        for item in finaldict.keys():
            finaldict[item].append(
                imagery_approval(file_cleanup("/Users/colingreenman/%s" % (item)))
            )
            finaldict[item].append(ethnicity())
            print(finaldict)
    printer(finaldict, "IG")


def engagement_rate():
    ig = instaloader.Instaloader()
    ig.load_session_from_file(universal_username)
    # for item in finaldict.keys():
    usrname = "cpgreenman"
    profile = instaloader.Profile.from_username(ig.context, usrname)

    print(profile.username)
    followers = profile.followers
    print(followers)
    print(profile.biography)
    # finaldict[item].append(followers)
    count = 0
    engagement_count = 0
    for post in profile.get_posts():
        if count < 100:
            count += 1
            engagement_count += int(post.likes)
            engagement_count += int(post.comments)


def similar_account(alreadylist=[]):
    username = input("Username to search similar accounts for: ")
    ig = instaloader.Instaloader()
    # ig.login("sylk_ny","Raindance95!")
    # ig.test_login()
    ig.load_session_from_file(universal_username)
    prof = instaloader.Profile.from_username(ig.context, username)
    rng_bottom = int(input("Follower Range Minimum: "))
    rng_top = int(input("Follower Range Max: "))
    count = 1
    approval_lst = []
    if len(alreadylist) < 1:
        driver = webdriver.Chrome()
    """driver.get("https://instagram.com/")
	time.sleep(2)
	text_box = driver.find_element(By.CSS_SELECTOR, '[name="username"]')
	text_box.send_keys("sylk_ny");
	time.sleep(2)
	text_box = driver.find_element(By.CSS_SELECTOR, '[name="password"]')
	text_box.send_keys("Raindance95!");
	time.sleep(2)
	text_box.send_keys(Keys.RETURN)
	#time.sleep(2)
	#continue_link = driver.find_element(By.LINK_TEXT, 'Log in')
	#continue_link.click()"""
    x = input("Have you logged in in the AUTO-WINDOW? (PRESS ANY KEY)")
    count = 0
    for act in prof.get_similar_accounts():
        count += 1
        usr = act.username
        print(count, str(act), usr)
        if usr in alreadylist:
            continue
        alreadylist.append(usr)
        r = requests.get("https://instagram.com/%s" % act.username)
        soup = BeautifulSoup(r.text, "html.parser")
        base = soup.find(property="og:description")
        followers = str(base)[15:19]
        if "K" in followers:
            followers = followers[: followers.find("K")]
            followers = int(followers) * 1000
        if "M" in str(followers):
            followers = followers[: followers.find("M")]
            followers = int(followers) * 1000000
        if "," in str(followers):
            temp = str(followers).find(",")
            followers = str(followers)[:temp] + str(followers)[temp + 1 :]
            # print ("comma removed",temp,str(followers))
            followers = int(followers) * 10

        print(followers)
        print("------------------------")
        if "F" in (str(followers))[-1] or "o" in (str(followers))[-1]:
            continue
        if int(followers) < rng_top * 1.1 and int(followers) > rng_bottom * 0.9:
            print(act.username, followers, "!!!!!!")
            driver.get("https://instagram.com/%s/" % act.username)
            choice = input("Add this Creator? (1=Yes)")
            if choice == "1":
                approval_lst.append("https://instagram.com/%s/" % act.username)
            continue
    for item in approval_lst:
        print(item)
    new_usr = input(
        "Would you like to re-run this program with a new influencer? 1 = Yes"
    )
    if new_usr == "1":
        similar_account(alreadylist)


similar_account()
exit()

data_pull(IG_usernames())
# printer(finaldict)
# engagement_rate()
exit()
data_pull(usernames())
printer(finaldict)
print(finaldict)
exit()
folder = imagery_approval(file_cleanup("/Users/colingreenman/%s" % (profile.username)))
